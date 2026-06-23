from __future__ import annotations

import hashlib
import importlib.util
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Callable

import fitz
import streamlit as st


APP_TITLE = "Resolution Law Tools"
FOOTER_TOP_RATIO = 0.92
OFN_FOOTER_TOP_RATIO = 0.65
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
OFN_PATTERN = re.compile(r"\b[O0]\s*F\s*N\s*[:;#-]?\s*([0-9][0-9\s,.\-]{1,24})", re.IGNORECASE)


@dataclass
class ToolDefinition:
    tool_id: str
    name: str
    category: str
    description: str
    render: Callable[[], None]


@dataclass
class ChunkRow:
    row_id: str
    source_file: str
    pages: str
    start_page: int
    end_page: int
    detected_beam: str
    filename_stem: str
    method: str
    issue: str
    include: bool


@dataclass
class SplitResult:
    source_file: str
    pages: str
    output_file: str
    beam_number: str
    method: str
    status: str


@dataclass
class Splitter2Result:
    source_file: str
    pages: str
    output_file: str
    ofn_number: str
    method: str
    status: str


def inject_css() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background:
                linear-gradient(180deg, #f8fafc 0%, #eef4ff 44%, #f8fafc 100%);
            color: #0f172a;
        }
        [data-testid="stHeader"] {
            background: rgba(248, 250, 252, 0.88);
        }
        [data-testid="stSidebar"] {
            background: #0f172a;
        }
        [data-testid="stSidebar"] * {
            color: #e5e7eb;
        }
        [data-testid="stSidebar"] div[role="radiogroup"] label {
            border-radius: 12px;
            padding: 0.25rem 0.35rem;
        }
        .hub-hero {
            background: #ffffff;
            color: #0f172a;
            padding: 2rem 2.15rem;
            border: 1px solid #dbe4f0;
            border-radius: 20px;
            margin-bottom: 1.25rem;
            box-shadow: 0 18px 45px rgba(15, 23, 42, 0.08);
            position: relative;
            overflow: hidden;
        }
        .hub-hero:before {
            content: "";
            position: absolute;
            inset: 0 0 auto 0;
            height: 5px;
            background: linear-gradient(90deg, #2563eb, #14b8a6, #facc15);
        }
        .hub-hero h1 {
            margin: 0;
            font-size: 2.35rem;
            line-height: 1.05;
            letter-spacing: 0;
            color: #0f172a;
        }
        .hub-hero p {
            margin: 0.75rem 0 0;
            color: #475569;
            font-size: 1.03rem;
            max-width: 760px;
        }
        .tool-card {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 16px;
            padding: 1.15rem 1.2rem;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.05);
            min-height: 134px;
            transition: border-color 140ms ease, box-shadow 140ms ease, transform 140ms ease;
        }
        .tool-card.clickable {
            min-height: 160px;
        }
        .tool-card.clickable:hover {
            border-color: #93c5fd;
            box-shadow: 0 18px 40px rgba(37, 99, 235, 0.12);
            transform: translateY(-1px);
        }
        .tool-card h3 {
            margin: 0 0 0.35rem;
            color: #0f172a;
        }
        .tool-card p,
        .small-muted {
            color: #667085;
            font-size: 0.92rem;
        }
        .metric-strip {
            display: grid;
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap: 0.75rem;
            margin: 1rem 0;
        }
        .metric-box {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 14px;
            padding: 0.85rem 1rem;
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.04);
        }
        .metric-box strong {
            display: block;
            color: #0f172a;
            font-size: 1.45rem;
        }
        .metric-box span {
            color: #667085;
            font-size: 0.88rem;
        }
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stButton"] button {
            border-radius: 12px;
            font-weight: 650;
            border: 1px solid #cbd5e1;
            box-shadow: 0 5px 14px rgba(15, 23, 42, 0.06);
        }
        .section-label {
            color: #0f172a;
            font-size: 1.1rem;
            font-weight: 700;
            margin: 1rem 0 0.25rem;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.5rem;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius: 999px;
            background: #ffffff;
            border: 1px solid #e2e8f0;
            padding: 0.45rem 0.85rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def find_tesseract_executable() -> Path | None:
    env_value = os.environ.get("TESSERACT_CMD")
    if env_value and Path(env_value).exists():
        return Path(env_value)

    path_value = shutil.which("tesseract")
    if path_value:
        return Path(path_value)

    if sys.platform.startswith("win"):
        candidates = [
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Tesseract-OCR" / "tesseract.exe",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Tesseract-OCR" / "tesseract.exe",
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate

    return None


def dependency_status() -> dict[str, bool]:
    return {
        "PyMuPDF": importlib.util.find_spec("fitz") is not None,
        "Pillow": importlib.util.find_spec("PIL") is not None,
        "pytesseract": importlib.util.find_spec("pytesseract") is not None,
        "Tesseract OCR": find_tesseract_executable() is not None,
    }


def ocr_available() -> bool:
    return (
        importlib.util.find_spec("pytesseract") is not None
        and importlib.util.find_spec("PIL") is not None
        and find_tesseract_executable() is not None
    )


def configure_tesseract(pytesseract_module) -> None:
    executable = find_tesseract_executable()
    if executable is not None:
        pytesseract_module.pytesseract.tesseract_cmd = str(executable)


def sanitize_filename_stem(value: str, fallback: str) -> str:
    stem = value.strip()
    if stem.lower().endswith(".pdf"):
        stem = stem[:-4]
    stem = INVALID_FILENAME_CHARS.sub("_", stem)
    stem = re.sub(r"\s+", "_", stem).strip(" ._")
    return stem or fallback


def make_unique_stem(existing: set[str], requested_stem: str, fallback: str) -> str:
    base = sanitize_filename_stem(requested_stem, fallback)
    candidate = base
    counter = 2
    while candidate.lower() in existing:
        candidate = f"{base}_{counter:02d}"
        counter += 1
    existing.add(candidate.lower())
    return candidate


def number_tokens(text: str) -> list[str]:
    if not text:
        return []

    normalized = text.replace(",", "")
    tokens = re.findall(r"\b\d{2,}\b", normalized)
    if tokens:
        return tokens

    digits_only = re.sub(r"\D", "", normalized)
    if len(digits_only) >= 2:
        return [digits_only]

    return []


def find_beam_number_in_text(text: str) -> str | None:
    tokens = number_tokens(text)
    return tokens[-1] if tokens else None


def find_native_footer_number(page) -> str | None:
    page_rect = page.rect
    footer_y = page_rect.height * 0.74
    candidates: list[tuple[float, float, str]] = []

    for word in page.get_text("words"):
        x0, y0, x1, y1, text = word[:5]
        if y0 < footer_y:
            continue
        for token in number_tokens(text):
            candidates.append((float(y1), float(x0), token))

    if not candidates:
        text = page.get_text("text", clip=fitz.Rect(0, footer_y, page_rect.width, page_rect.height)) or ""
        return find_beam_number_in_text(text)

    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[-1][2]


def find_ocr_footer_number(page, ocr_error_callback: Callable[[Exception], None] | None = None) -> tuple[str | None, str]:
    try:
        from PIL import Image, ImageOps
        import pytesseract

        configure_tesseract(pytesseract)
        page_rect = page.rect
        footer_rect = fitz.Rect(0, page_rect.height * 0.74, page_rect.width, page_rect.height)
        matrix = fitz.Matrix(3, 3)
        pixmap = page.get_pixmap(matrix=matrix, clip=footer_rect, alpha=False)
        image = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
        image = ImageOps.autocontrast(ImageOps.grayscale(image))

        data = pytesseract.image_to_data(
            image,
            config="--psm 6 -c tessedit_char_whitelist=0123456789",
            output_type=pytesseract.Output.DICT,
        )

        candidates: list[tuple[float, float, str]] = []
        for index, text in enumerate(data.get("text", [])):
            for token in number_tokens(text):
                top = float(data["top"][index])
                height = float(data["height"][index])
                left = float(data["left"][index])
                candidates.append((top + height, left, token))

        if candidates:
            candidates.sort(key=lambda item: (item[0], item[1]))
            return candidates[-1][2], "OCR"

        ocr_text = pytesseract.image_to_string(
            image,
            config="--psm 6 -c tessedit_char_whitelist=0123456789",
        )
        beam = find_beam_number_in_text(ocr_text)
        if beam:
            return beam, "OCR"
    except Exception as exc:
        if ocr_error_callback is not None:
            ocr_error_callback(exc)

    return None, "not found"


def extract_beam_number_from_page(
    page,
    use_ocr: bool,
    ocr_error_callback: Callable[[Exception], None] | None = None,
) -> tuple[str | None, str]:
    beam = find_native_footer_number(page)
    if beam:
        return beam, "native text"

    if not use_ocr:
        return None, "not found"

    return find_ocr_footer_number(page, ocr_error_callback)


def normalize_ofn_number(value: str) -> str | None:
    digits = re.sub(r"\D", "", value or "")
    return digits if len(digits) >= 2 else None


def find_ofn_number_in_text(text: str) -> str | None:
    if not text:
        return None

    normalized = " ".join(text.split())
    for match in OFN_PATTERN.finditer(normalized):
        ofn_number = normalize_ofn_number(match.group(1))
        if ofn_number:
            return ofn_number

    return None


def find_ofn_number_in_tokens(tokens: list[str]) -> str | None:
    clean_tokens = [token.strip() for token in tokens if token and token.strip()]
    for index in range(len(clean_tokens)):
        candidate_text = " ".join(clean_tokens[index:index + 6])
        ofn_number = find_ofn_number_in_text(candidate_text)
        if ofn_number:
            return ofn_number

    return None


def find_native_footer_ofn_number(page) -> str | None:
    page_rect = page.rect
    footer_y = page_rect.height * OFN_FOOTER_TOP_RATIO
    footer_rect = fitz.Rect(0, footer_y, page_rect.width, page_rect.height)
    text = page.get_text("text", clip=footer_rect) or ""
    ofn_number = find_ofn_number_in_text(text)
    if ofn_number:
        return ofn_number

    words = [
        str(word[4])
        for word in page.get_text("words")
        if len(word) >= 5 and float(word[1]) >= footer_y
    ]
    return find_ofn_number_in_tokens(words)


def find_ocr_footer_ofn_number(page, ocr_error_callback: Callable[[Exception], None] | None = None) -> tuple[str | None, str]:
    try:
        from PIL import Image, ImageOps
        import pytesseract

        configure_tesseract(pytesseract)
        page_rect = page.rect
        footer_rect = fitz.Rect(0, page_rect.height * OFN_FOOTER_TOP_RATIO, page_rect.width, page_rect.height)
        matrix = fitz.Matrix(3, 3)
        pixmap = page.get_pixmap(matrix=matrix, clip=footer_rect, alpha=False)
        image = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
        image = ImageOps.autocontrast(ImageOps.grayscale(image))

        for config in ("--psm 6", "--psm 11"):
            ocr_text = pytesseract.image_to_string(image, config=config)
            ofn_number = find_ofn_number_in_text(ocr_text)
            if ofn_number:
                return ofn_number, "OCR"

        data = pytesseract.image_to_data(
            image,
            config="--psm 6",
            output_type=pytesseract.Output.DICT,
        )
        ofn_number = find_ofn_number_in_tokens(data.get("text", []))
        if ofn_number:
            return ofn_number, "OCR"
    except Exception as exc:
        if ocr_error_callback is not None:
            ocr_error_callback(exc)

    return None, "not found"


def extract_ofn_number_from_page(
    page,
    use_ocr: bool,
    ocr_error_callback: Callable[[Exception], None] | None = None,
) -> tuple[str | None, str]:
    ofn_number = find_native_footer_ofn_number(page)
    if ofn_number:
        return ofn_number, "native text"

    if not use_ocr:
        return None, "not found"

    return find_ocr_footer_ofn_number(page, ocr_error_callback)


def save_page_range_to_bytes(doc, start_page: int, end_page: int) -> bytes:
    out_doc = fitz.open()
    try:
        out_doc.insert_pdf(doc, from_page=start_page, to_page=end_page)
        buffer = BytesIO()
        out_doc.save(buffer, garbage=4, deflate=True)
        return buffer.getvalue()
    finally:
        out_doc.close()


def file_hash(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()[:12]


def uploaded_files_signature(uploaded_files) -> str:
    parts: list[str] = []
    for uploaded_file in uploaded_files:
        data = uploaded_file.getvalue()
        parts.append(f"{uploaded_file.name}:{len(data)}:{file_hash(data)}")
    return "|".join(sorted(parts))


def analyze_uploads(uploaded_files, odd_page_mode: str, missing_beam_mode: str) -> tuple[list[dict], list[str], dict[str, bytes]]:
    rows: list[ChunkRow] = []
    logs: list[str] = []
    file_bytes: dict[str, bytes] = {}
    fallback_counter = 1
    use_ocr = ocr_available()
    ocr_warning_logged = False

    if not use_ocr:
        logs.append("OCR is unavailable. Native PDFs can still be processed, but scanned PDFs need Tesseract OCR on the server.")

    def warn_ocr_once(exc: Exception) -> None:
        nonlocal ocr_warning_logged
        if not ocr_warning_logged:
            logs.append(f"OCR warning: {exc}")
            ocr_warning_logged = True

    for uploaded_file in uploaded_files:
        data = uploaded_file.getvalue()
        source_name = Path(uploaded_file.name).name
        source_id = f"{file_hash(data)}_{source_name}"
        file_bytes[source_id] = data

        try:
            doc = fitz.open(stream=data, filetype="pdf")
        except Exception as exc:
            logs.append(f"{source_name}: could not open PDF ({exc}).")
            continue

        try:
            page_count = doc.page_count
            logs.append(f"{source_name}: opened {page_count} page(s).")
            paired_page_count = page_count - (page_count % 2)

            for start_page in range(0, paired_page_count, 2):
                beam_number, method = extract_beam_number_from_page(doc[start_page + 1], use_ocr, warn_ocr_once)
                issue = ""
                include = True
                filename_stem = beam_number or ""

                if not beam_number:
                    if missing_beam_mode == "Skip missing beam pairs":
                        include = False
                        issue = "No beam number found; skipped by rule."
                    elif missing_beam_mode == "Use fallback names":
                        filename_stem = f"unnamed_{fallback_counter:03d}"
                        fallback_counter += 1
                        issue = "No beam number found; fallback name assigned."
                    else:
                        issue = "No beam number found; enter filename or uncheck include."

                rows.append(
                    ChunkRow(
                        row_id=f"{source_id}:{start_page}:{start_page + 1}",
                        source_file=source_name,
                        pages=f"{start_page + 1}-{start_page + 2}",
                        start_page=start_page,
                        end_page=start_page + 1,
                        detected_beam=beam_number or "",
                        filename_stem=filename_stem,
                        method=method,
                        issue=issue,
                        include=include,
                    )
                )

            if page_count % 2:
                orphan_page = page_count - 1
                if odd_page_mode == "Skip odd final pages":
                    filename_stem = ""
                    issue = "Odd final page; skipped by rule."
                    include = False
                elif odd_page_mode == "Review odd final pages":
                    filename_stem = ""
                    issue = "Odd final page; enter filename or uncheck include."
                    include = True
                else:
                    filename_stem = f"{Path(source_name).stem}_page_{page_count:03d}"
                    issue = "Odd final page saved as single-page output."
                    include = True

                rows.append(
                    ChunkRow(
                        row_id=f"{source_id}:{orphan_page}:{orphan_page}",
                        source_file=source_name,
                        pages=str(page_count),
                        start_page=orphan_page,
                        end_page=orphan_page,
                        detected_beam="",
                        filename_stem=filename_stem,
                        method="odd page",
                        issue=issue,
                        include=include,
                    )
                )
        finally:
            doc.close()

    return [asdict(row) for row in rows], logs, file_bytes


def build_zip(rows: list[dict], file_bytes: dict[str, bytes]) -> tuple[bytes | None, list[str]]:
    errors: list[str] = []
    output = BytesIO()
    used_names: set[str] = set()
    written = 0

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for row in rows:
            if not row.get("include"):
                continue

            filename_stem = str(row.get("filename_stem") or "").strip()
            if not filename_stem:
                errors.append(f"{row['source_file']} pages {row['pages']}: filename is blank.")
                continue

            source_key = str(row["row_id"]).split(":", 1)[0]
            data = file_bytes.get(source_key)
            if data is None:
                errors.append(f"{row['source_file']} pages {row['pages']}: source file data was not found.")
                continue

            try:
                doc = fitz.open(stream=data, filetype="pdf")
                try:
                    pdf_bytes = save_page_range_to_bytes(doc, int(row["start_page"]), int(row["end_page"]))
                finally:
                    doc.close()
            except Exception as exc:
                errors.append(f"{row['source_file']} pages {row['pages']}: could not split PDF ({exc}).")
                continue

            unique_stem = make_unique_stem(used_names, filename_stem, f"unnamed_{written + 1:03d}")
            archive.writestr(f"beam_split_results/{unique_stem}.pdf", pdf_bytes)
            written += 1

    if errors:
        return None, errors
    if written == 0:
        return None, ["No output files were selected. Check at least one row to include."]

    return output.getvalue(), []


def split_and_rename_pdfs(uploaded_files, include_odd_final_page: bool, use_fallback_names: bool) -> tuple[bytes | None, list[dict], list[str]]:
    logs: list[str] = []
    rows: list[SplitResult] = []
    output = BytesIO()
    used_names: set[str] = set()
    fallback_counter = 1
    written = 0
    use_ocr = ocr_available()
    ocr_warning_logged = False

    if not use_ocr:
        logs.append("OCR is unavailable on the server. Native PDFs may still work, but scanned PDFs need Tesseract OCR.")

    def warn_ocr_once(exc: Exception) -> None:
        nonlocal ocr_warning_logged
        if not ocr_warning_logged:
            logs.append(f"OCR warning: {exc}")
            ocr_warning_logged = True

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for uploaded_file in uploaded_files:
            data = uploaded_file.getvalue()
            source_name = Path(uploaded_file.name).name

            try:
                doc = fitz.open(stream=data, filetype="pdf")
            except Exception as exc:
                logs.append(f"{source_name}: could not open PDF ({exc}).")
                continue

            try:
                page_count = doc.page_count
                pair_page_count = page_count - (page_count % 2)

                for start_page in range(0, pair_page_count, 2):
                    end_page = start_page + 1
                    pages = f"{start_page + 1}-{end_page + 1}"
                    beam_number, method = extract_beam_number_from_page(doc[end_page], use_ocr, warn_ocr_once)
                    status = "Beam number found."

                    if beam_number:
                        stem = beam_number
                    elif use_fallback_names:
                        stem = f"unnamed_{fallback_counter:03d}"
                        fallback_counter += 1
                        status = "Beam number not found; fallback name used."
                    else:
                        logs.append(f"{source_name} pages {pages}: beam number not found; skipped.")
                        rows.append(
                            SplitResult(
                                source_file=source_name,
                                pages=pages,
                                output_file="",
                                beam_number="",
                                method="not found",
                                status="Skipped - beam number not found.",
                            )
                        )
                        continue

                    unique_stem = make_unique_stem(used_names, stem, f"unnamed_{fallback_counter:03d}")
                    output_file = f"{unique_stem}.pdf"
                    pdf_bytes = save_page_range_to_bytes(doc, start_page, end_page)
                    archive.writestr(output_file, pdf_bytes)
                    written += 1
                    rows.append(
                        SplitResult(
                            source_file=source_name,
                            pages=pages,
                            output_file=output_file,
                            beam_number=beam_number or "",
                            method=method,
                            status=status,
                        )
                    )

                if page_count % 2:
                    last_page = page_count - 1
                    if include_odd_final_page:
                        stem = f"{Path(source_name).stem}_page_{page_count:03d}"
                        unique_stem = make_unique_stem(used_names, stem, stem)
                        output_file = f"{unique_stem}.pdf"
                        pdf_bytes = save_page_range_to_bytes(doc, last_page, last_page)
                        archive.writestr(output_file, pdf_bytes)
                        written += 1
                        rows.append(
                            SplitResult(
                                source_file=source_name,
                                pages=str(page_count),
                                output_file=output_file,
                                beam_number="",
                                method="odd final page",
                                status="Included as single-page output.",
                            )
                        )
                    else:
                        logs.append(f"{source_name} page {page_count}: odd final page skipped.")
            finally:
                doc.close()

    if written == 0:
        return None, [asdict(row) for row in rows], logs + ["No PDF chunks were created."]

    return output.getvalue(), [asdict(row) for row in rows], logs


# ============================================================================
# Affidavit PDF Splitter
# Splits affidavit PDFs into two-page chunks and names outputs from OFN numbers.
# ============================================================================
def split_and_rename_ofn_pdfs(uploaded_files, include_odd_final_page: bool, use_fallback_names: bool) -> tuple[bytes | None, list[dict], list[str]]:
    logs: list[str] = []
    rows: list[Splitter2Result] = []
    output = BytesIO()
    used_names: set[str] = set()
    fallback_counter = 1
    written = 0
    use_ocr = ocr_available()
    ocr_warning_logged = False

    if not use_ocr:
        logs.append("OCR is unavailable on the server. Native PDFs may still work, but scanned PDFs need Tesseract OCR.")

    def warn_ocr_once(exc: Exception) -> None:
        nonlocal ocr_warning_logged
        if not ocr_warning_logged:
            logs.append(f"OCR warning: {exc}")
            ocr_warning_logged = True

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for uploaded_file in uploaded_files:
            data = uploaded_file.getvalue()
            source_name = Path(uploaded_file.name).name

            try:
                doc = fitz.open(stream=data, filetype="pdf")
            except Exception as exc:
                logs.append(f"{source_name}: could not open PDF ({exc}).")
                continue

            try:
                page_count = doc.page_count
                pair_page_count = page_count - (page_count % 2)

                for start_page in range(0, pair_page_count, 2):
                    end_page = start_page + 1
                    pages = f"{start_page + 1}-{end_page + 1}"
                    ofn_number, method = extract_ofn_number_from_page(doc[end_page], use_ocr, warn_ocr_once)
                    status = "OFN number found."

                    if ofn_number:
                        stem = f"{ofn_number}W"
                    elif use_fallback_names:
                        stem = f"unnamed_{fallback_counter:03d}"
                        fallback_counter += 1
                        status = "OFN number not found; fallback name used."
                    else:
                        logs.append(f"{source_name} pages {pages}: OFN number not found; skipped.")
                        rows.append(
                            Splitter2Result(
                                source_file=source_name,
                                pages=pages,
                                output_file="",
                                ofn_number="",
                                method="not found",
                                status="Skipped - OFN number not found.",
                            )
                        )
                        continue

                    unique_stem = make_unique_stem(used_names, stem, f"unnamed_{fallback_counter:03d}")
                    output_file = f"{unique_stem}.pdf"
                    pdf_bytes = save_page_range_to_bytes(doc, start_page, end_page)
                    archive.writestr(output_file, pdf_bytes)
                    written += 1
                    rows.append(
                        Splitter2Result(
                            source_file=source_name,
                            pages=pages,
                            output_file=output_file,
                            ofn_number=ofn_number or "",
                            method=method,
                            status=status,
                        )
                    )

                if page_count % 2:
                    last_page = page_count - 1
                    if include_odd_final_page:
                        stem = f"{Path(source_name).stem}_page_{page_count:03d}"
                        unique_stem = make_unique_stem(used_names, stem, stem)
                        output_file = f"{unique_stem}.pdf"
                        pdf_bytes = save_page_range_to_bytes(doc, last_page, last_page)
                        archive.writestr(output_file, pdf_bytes)
                        written += 1
                        rows.append(
                            Splitter2Result(
                                source_file=source_name,
                                pages=str(page_count),
                                output_file=output_file,
                                ofn_number="",
                                method="odd final page",
                                status="Included as single-page output.",
                            )
                        )
                    else:
                        logs.append(f"{source_name} page {page_count}: odd final page skipped.")
            finally:
                doc.close()

    if written == 0:
        return None, [asdict(row) for row in rows], logs + ["No PDF chunks were created."]

    return output.getvalue(), [asdict(row) for row in rows], logs


# ============================================================================
# Spreadsheet Comparison Tool — Resolution Law Tools
# Compares two snapshots of the same tracker spreadsheet (.xlsx or .csv).
# User picks one or more key columns (composite key supported for cases where
# a single column isn't unique per row, e.g. Beam Number + Garnishee when one
# case has multiple garnishments). Reports added rows, removed rows, and per-
# cell changes after normalizing away formatting noise.
# ============================================================================
import csv as _csv
import datetime as _dt
import io as _io
import re as _re
from collections import Counter as _Counter

import openpyxl as _openpyxl
from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment

_DATE_RE    = _re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")
_TIME_RE    = _re.compile(r"^(\d{1,2}):(\d{2}):(\d{2})\s*(AM|PM|am|pm)?$")
_MONEY_RE   = _re.compile(r"^\(\s*-?\$?\s*-?[\d,]+(?:\.\d+)?\s*\)$|^-?\$\s*-?[\d,]+(?:\.\d+)?$")
_ILLEGAL_RE = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

_KEY_HINTS = [
    "Beam Number", "Beam #", "Beam",
    "Case Number", "Case #", "Case No", "Case",
    "Account ID", "Account_ID", "Client Account ID", "Account",
    "Matter Number", "Matter #", "Matter ID", "Matter",
    "File Number", "File #", "File ID",
    "ID", "Id",
]

# Substrings that mark a column as "volatile" — its value changes across snapshots,
# so it shouldn't be part of an auto-detected composite key. (User can still pick
# manually if they want.)
_KEY_ANTI_HINTS = [
    "answer", "status", "amount", "holding", "cost", "balance", "paid",
    "due", "collected", "notes", "comment", "remark", "received",
    "withheld", "withholding", "response", "result", "outcome", "phase",
    "stage", "stop", "motion", "order",
]


def _norm_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, _dt.date):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v != v:
            return ""
        if v.is_integer():
            return str(int(v))
        return f"{v:.6f}".rstrip("0").rstrip(".")
    s = str(v).replace(" ", " ")
    s = _ILLEGAL_RE.sub("", s).strip()
    s = _re.sub(r"\s+", " ", s)
    if s == "":
        return ""
    m = _DATE_RE.match(s)
    if m:
        mo, da, yr = m.groups()
        return f"{int(mo):02d}/{int(da):02d}/{yr}"
    m = _TIME_RE.match(s)
    if m:
        hh, mm, ss, ap = m.groups()
        suf = f" {ap.upper()}" if ap else ""
        return f"{int(hh):02d}:{mm}:{ss}{suf}"
    if _MONEY_RE.match(s):
        raw = s.replace(",", "")
        neg = False
        if raw.startswith("(") and raw.endswith(")"):
            neg = True; raw = raw[1:-1].strip()
        if raw.startswith("-"):
            neg = not neg; raw = raw[1:].strip()
        if raw.startswith("$"):
            raw = raw[1:].strip()
        if raw.startswith("-"):
            neg = not neg; raw = raw[1:].strip()
        try:
            val = float(raw)
            if neg: val = -val
            return f"${val:.2f}"
        except ValueError:
            pass
    try:
        f = float(s.replace(",", ""))
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


def _display_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return f"{v.month:02d}/{v.day:02d}/{v.year}"
        return v.strftime("%m/%d/%Y %H:%M:%S")
    if isinstance(v, _dt.date):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    if isinstance(v, float):
        if v != v: return ""
        if v.is_integer(): return str(int(v))
        return f"{v}"
    s = str(v)
    s = _ILLEGAL_RE.sub("", s)
    return s


def _detect_header_row(rows):
    for i, r in enumerate(rows):
        if any(c is not None and str(c).strip() != "" for c in r):
            return i
    return 0


def _read_xlsx_bytes(data: bytes) -> tuple[str, list[str], list[dict]]:
    wb = _openpyxl.load_workbook(_io.BytesIO(data), data_only=True)
    ws = wb.active
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not all_rows:
        return ws.title, [], []
    header_idx = _detect_header_row(all_rows)
    raw = all_rows[header_idx]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for c in raw:
        h = "" if c is None else str(c).strip()
        if h == "":
            headers.append(""); continue
        if h in seen:
            seen[h] += 1
            headers.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            headers.append(h)
    rows = []
    for r in all_rows[header_idx + 1:]:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in r):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h: continue
            rec[h] = r[i] if i < len(r) else None
        rows.append(rec)
    return ws.title, [h for h in headers if h], rows


def _read_csv_bytes(data: bytes) -> tuple[str, list[str], list[dict]]:
    try:
        text = data.decode("utf-8-sig", errors="replace")
    except Exception:
        text = data.decode("latin-1", errors="replace")
    reader = _csv.reader(_io.StringIO(text))
    all_rows = [list(r) for r in reader]
    if not all_rows:
        return "csv", [], []
    header_idx = _detect_header_row(all_rows)
    raw = all_rows[header_idx]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for c in raw:
        h = "" if c is None else str(c).strip()
        if h == "":
            headers.append(""); continue
        if h in seen:
            seen[h] += 1
            headers.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            headers.append(h)
    rows = []
    for r in all_rows[header_idx + 1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h: continue
            rec[h] = r[i] if i < len(r) else None
        rows.append(rec)
    return "csv", [h for h in headers if h], rows


def _read_uploaded(uploaded) -> tuple[str, list[str], list[dict]]:
    name = uploaded.name.lower()
    data = uploaded.getvalue()
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _read_csv_bytes(data)
    return _read_xlsx_bytes(data)


def _suggest_key(headers_a, headers_b) -> str | None:
    common = [h for h in headers_a if h in set(headers_b)]
    common_lower = {h.lower(): h for h in common}
    for hint in _KEY_HINTS:
        if hint.lower() in common_lower:
            return common_lower[hint.lower()]
    return common[0] if common else None


def _row_key(rec, key_cols):
    return "\t".join(_norm_value(rec.get(c)) for c in key_cols)


def _is_volatile_col(col_name: str) -> bool:
    """Columns whose names suggest they hold value/state data that changes
    between snapshots — bad candidates for an auto-detected key."""
    lower = col_name.lower()
    return any(anti in lower for anti in _KEY_ANTI_HINTS)


def _smart_suggest_keys(rows_a, rows_b, common_cols, max_cols=4):
    """Greedy auto-detect a composite key. Starts with the heuristic single-column
    suggestion and extends with whichever extra column reduces the duplicate-key
    count the most, until duplicates hit zero or no column helps. Capped at max_cols.

    The extender prefers "identifying" columns (names, dates, IDs) and avoids
    "volatile" columns (Answer, Status, Amount, etc.) — those change between
    snapshots, so including them in the key would make a single modified row
    look like one removed + one new."""
    if not common_cols:
        return []
    seed = _suggest_key(common_cols, common_cols)
    if seed not in common_cols:
        seed = common_cols[0]
    chosen = [seed]

    def dup_count(cols):
        keys_a = [_row_key(r, cols) for r in rows_a]
        keys_b = [_row_key(r, cols) for r in rows_b]
        ca = _Counter(keys_a); cb = _Counter(keys_b)
        return sum(1 for c in ca.values() if c > 1) + sum(1 for c in cb.values() if c > 1)

    # First pass: only consider non-volatile columns
    safe_cols = [c for c in common_cols if not _is_volatile_col(c)]
    while len(chosen) < max_cols:
        current = dup_count(chosen)
        if current == 0:
            break
        best_col = None
        best_dups = current
        for c in safe_cols:
            if c in chosen:
                continue
            d = dup_count(chosen + [c])
            if d < best_dups:
                best_dups = d
                best_col = c
        if best_col is None:
            break
        chosen.append(best_col)
    return chosen


def _compare_rows(rows_a, rows_b, key_cols, compare_cols):
    """Return (new_rows, mod_summary, mod_details, rem_rows). key_cols supports composite keys."""
    by_a: dict[str, dict] = {}
    by_b: dict[str, dict] = {}
    for r in rows_a:
        k = _row_key(r, key_cols)
        if k.replace("\t", "") != "":
            by_a.setdefault(k, r)
    for r in rows_b:
        k = _row_key(r, key_cols)
        if k.replace("\t", "") != "":
            by_b.setdefault(k, r)

    sa, sb = set(by_a), set(by_b)
    new_keys = sorted(sb - sa)
    rem_keys = sorted(sa - sb)
    com_keys = sorted(sa & sb)
    key_set  = set(key_cols)

    def _key_rec(k):
        parts = k.split("\t")
        return {key_cols[i]: parts[i] if i < len(parts) else "" for i in range(len(key_cols))}

    new_out = []
    for k in new_keys:
        rec = _key_rec(k)
        for c in compare_cols:
            if c in key_set: continue
            rec[c] = _display_value(by_b[k].get(c))
        new_out.append(rec)

    rem_out = []
    for k in rem_keys:
        rec = _key_rec(k)
        for c in compare_cols:
            if c in key_set: continue
            rec[c] = _display_value(by_a[k].get(c))
        rem_out.append(rec)

    mod_summary: list[dict] = []
    mod_details: list[dict] = []
    for k in com_keys:
        ra = by_a[k]; rb = by_b[k]
        key_rec = _key_rec(k)
        changed: list[str] = []
        for c in compare_cols:
            if c in key_set: continue
            va = _norm_value(ra.get(c))
            vb = _norm_value(rb.get(c))
            if va != vb:
                changed.append(c)
                mod_details.append({
                    **key_rec,
                    "Column": c,
                    "Old Value": _display_value(ra.get(c)),
                    "New Value": _display_value(rb.get(c)),
                })
        if changed:
            mod_summary.append({
                **key_rec,
                "Columns Changed": ", ".join(changed),
                "Change Count": len(changed),
            })

    return new_out, mod_summary, mod_details, rem_out


def _build_compare_excel(new_rows, mod_summary, mod_details, rem_rows, compare_cols, key_cols) -> bytes:
    wb = _openpyxl.Workbook()

    def write(name, rows, headers, color, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.font = _Font(bold=True, color="FFFFFF")
            c.fill = _PatternFill("solid", fgColor=color)
            c.alignment = _Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for ci, h in enumerate(headers, 1):
            samples = [len(str(h))] + [min(len(str(r.get(h, ""))), 80) for r in rows[:300]]
            ml = max(samples) if samples else 12
            ws.column_dimensions[_openpyxl.utils.get_column_letter(ci)].width = min(max(12, ml + 2), 80)
        ws.freeze_panes = "A2"

    key_set = set(key_cols)
    other_cols = [c for c in compare_cols if c not in key_set]
    full_headers = list(key_cols) + other_cols
    summary_hdrs = list(key_cols) + ["Columns Changed", "Change Count"]
    details_hdrs = list(key_cols) + ["Column", "Old Value", "New Value"]

    write("New", new_rows, full_headers, "2E7D32", first=True)
    write("Modified (Summary)", mod_summary, summary_hdrs, "1565C0")
    write("Modified (Details)", mod_details, details_hdrs, "3949AB")
    write("No Longer In File", rem_rows, full_headers, "C62828")

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_spreadsheet_compare() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Spreadsheet Comparison</h1>
            <p>Upload two snapshots of the same spreadsheet (the older version and the newer version).
            The tool matches rows by a key column you pick, reports new, removed, and changed rows,
            and ignores formatting-only differences like date padding, time padding, dollar-sign
            spacing, parens-vs-minus, and CSV whitespace.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("How it works", expanded=False):
        st.markdown(
            "- Upload the **starting** spreadsheet (older) and the **ending** spreadsheet (newer).\n"
            "- The tool auto-detects a likely key column (Beam Number, Case Number, Account ID, etc.) "
            "and lets you adjust the selection.\n"
            "- If one column isn't unique per row (for example, a case has multiple garnishments), "
            "you can add more columns to make a composite key.\n"
            "- A row is flagged **Modified** only when one or more cell values differ after "
            "normalization (dates → `MM/DD/YYYY`, times → `HH:MM:SS AM/PM`, money → `$0.00` form, "
            "whitespace and quoting collapsed, control chars stripped).\n"
            "- Supports `.xlsx` and `.csv`. The first sheet of an `.xlsx` is used. Blank rows above "
            "the header are detected and skipped."
        )

    st.markdown('<div class="section-label">1. Upload the starting spreadsheet</div>', unsafe_allow_html=True)
    start_upload = st.file_uploader(
        "Starting spreadsheet (older snapshot)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=False,
        key="sc_start_file",
    )

    st.markdown('<div class="section-label">2. Upload the ending spreadsheet</div>', unsafe_allow_html=True)
    end_upload = st.file_uploader(
        "Ending spreadsheet (newer snapshot)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=False,
        key="sc_end_file",
    )

    if not (start_upload and end_upload):
        st.info("Upload both spreadsheets to continue.")
        return

    try:
        _sheet_a, headers_a, rows_a = _read_uploaded(start_upload)
    except Exception as exc:
        st.error(f"Could not read the starting spreadsheet: {exc}")
        return
    try:
        _sheet_b, headers_b, rows_b = _read_uploaded(end_upload)
    except Exception as exc:
        st.error(f"Could not read the ending spreadsheet: {exc}")
        return

    set_a, set_b = set(headers_a), set(headers_b)
    common_cols = [h for h in headers_a if h in set_b]
    only_a = [h for h in headers_a if h not in set_b]
    only_b = [h for h in headers_b if h not in set_a]

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(rows_a)}</strong><span>rows in starting file</span></div>
            <div class="metric-box"><strong>{len(rows_b)}</strong><span>rows in ending file</span></div>
            <div class="metric-box"><strong>{len(common_cols)}</strong><span>shared columns</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not common_cols:
        st.error(
            "The two spreadsheets have no shared columns, so rows can't be matched. "
            "Make sure both files have the same column headers."
        )
        if only_a:
            st.write(f"**Only in starting file:** {', '.join(only_a)}")
        if only_b:
            st.write(f"**Only in ending file:** {', '.join(only_b)}")
        return

    if only_a or only_b:
        with st.expander("Column differences (header-level)", expanded=False):
            if only_a:
                st.write(f"**Columns only in starting file:** {', '.join(only_a)}")
            if only_b:
                st.write(f"**Columns only in ending file:** {', '.join(only_b)}")
            st.caption("These columns are ignored when comparing rows. Only shared columns are compared.")

    default_keys = _smart_suggest_keys(rows_a, rows_b, common_cols)
    if not default_keys and common_cols:
        default_keys = [common_cols[0]]
    key_cols = st.multiselect(
        "Key column(s) — pick one or more to uniquely identify a row",
        options=common_cols,
        default=default_keys,
        help="Auto-detected composite key, chosen to minimize duplicate rows. Adjust if needed.",
        key="sc_key_cols",
    )
    if not key_cols:
        st.warning("Pick at least one key column.")
        return

    # Duplicate-key check on the chosen key
    keys_a_all = [_row_key(r, key_cols) for r in rows_a]
    keys_b_all = [_row_key(r, key_cols) for r in rows_b]
    dup_a = sum(1 for c in _Counter(keys_a_all).values() if c > 1)
    dup_b = sum(1 for c in _Counter(keys_b_all).values() if c > 1)
    if dup_a or dup_b:
        st.warning(
            f"Heads up: with the key column(s) you picked, **{dup_a}** duplicate key(s) exist in the "
            f"starting file and **{dup_b}** in the ending file. Only the FIRST row for each duplicate "
            "is compared. Add more columns to the key to make rows unique."
        )

    if st.button("Compare spreadsheets", type="primary", use_container_width=True):
        progress = st.progress(0, text="Comparing rows...")
        new_rows, mod_summary, mod_details, rem_rows = _compare_rows(
            rows_a, rows_b, key_cols, common_cols
        )
        progress.progress(80, text="Building Excel report...")
        excel_bytes = _build_compare_excel(
            new_rows, mod_summary, mod_details, rem_rows, common_cols, key_cols
        )
        progress.progress(100, text="Done")

        st.session_state["sc_new"]      = new_rows
        st.session_state["sc_summary"]  = mod_summary
        st.session_state["sc_details"]  = mod_details
        st.session_state["sc_rem"]      = rem_rows
        st.session_state["sc_excel"]    = excel_bytes
        st.session_state["sc_keys"]     = key_cols
        common_count = len(set(keys_a_all) & set(keys_b_all))
        st.session_state["sc_unchanged"] = max(0, common_count - len(mod_summary))

    if "sc_excel" in st.session_state:
        new_rows    = st.session_state["sc_new"]
        mod_summary = st.session_state["sc_summary"]
        mod_details = st.session_state["sc_details"]
        rem_rows    = st.session_state["sc_rem"]
        unchanged   = st.session_state.get("sc_unchanged", 0)

        st.markdown(
            f"""
            <div class="metric-strip">
                <div class="metric-box"><strong>{len(new_rows)}</strong><span>new rows</span></div>
                <div class="metric-box"><strong>{len(mod_summary)}</strong><span>modified rows ({len(mod_details)} cell changes)</span></div>
                <div class="metric-box"><strong>{len(rem_rows)}</strong><span>no longer in file</span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.caption(f"{unchanged} row(s) had no real changes after normalization (formatting-only differences ignored).")

        st.download_button(
            "Download Excel report",
            data=st.session_state["sc_excel"],
            file_name="Spreadsheet Comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        tab_new, tab_sum, tab_det, tab_rem = st.tabs(
            ["New", "Modified (Summary)", "Modified (Details)", "No Longer In File"]
        )
        with tab_new:
            if new_rows:
                st.dataframe(new_rows, use_container_width=True, hide_index=True)
            else:
                st.info("No new rows.")
        with tab_sum:
            if mod_summary:
                st.dataframe(mod_summary, use_container_width=True, hide_index=True)
            else:
                st.info("No modified rows.")
        with tab_det:
            if mod_details:
                st.dataframe(mod_details, use_container_width=True, hide_index=True)
            else:
                st.info("No cell-level changes.")
        with tab_rem:
            if rem_rows:
                st.dataframe(rem_rows, use_container_width=True, hide_index=True)
            else:
                st.info("No removed rows.")


def render_beam_pdf_splitter() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Beam PDF Splitter</h1>
            <p>Upload a PDF or a folder of PDFs, review every two-page chunk, fix any missing beam numbers, and download a clean ZIP of renamed files.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("Processing rules", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            include_odd_final_page = st.checkbox("Include odd final page as a single-page PDF", value=False)
        with col2:
            use_fallback_names = st.checkbox("Use fallback name if beam number is not found", value=False)

    st.markdown('<div class="section-label">Upload PDFs</div>', unsafe_allow_html=True)
    upload_tabs = st.tabs(["Single or multiple PDF files", "Folder of PDFs"])
    with upload_tabs[0]:
        file_uploads = st.file_uploader(
            "Choose one PDF or several PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            help="Use this when you have one PDF or a few PDFs selected manually.",
            key="beam_pdf_file_uploads",
        )
    with upload_tabs[1]:
        folder_uploads = st.file_uploader(
            "Choose a folder",
            type=["pdf"],
            accept_multiple_files="directory",
            help="Use this when you want to upload all PDFs from a folder.",
            key="beam_pdf_folder_uploads",
        )

    uploaded_files = list(file_uploads or []) + list(folder_uploads or [])
    upload_signature = uploaded_files_signature(uploaded_files)
    if st.session_state.get("beam_upload_signature") != upload_signature:
        st.session_state["beam_upload_signature"] = upload_signature
        st.session_state["beam_zip_bytes"] = None
        st.session_state["beam_result_rows"] = []
        st.session_state["beam_result_logs"] = []

    if not uploaded_files:
        left, right = st.columns(2)
        with left:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>What it does</h3>
                    <p>Splits page pairs like 1-2, 3-4, 5-6 and names each output from the beam number at the bottom of the second page.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with right:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>Review before download</h3>
                    <p>If OCR misses a beam number, edit the filename in the review table before creating the ZIP.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        return

    expected_pairs = 0
    for uploaded_file in uploaded_files:
        try:
            with fitz.open(stream=uploaded_file.getvalue(), filetype="pdf") as doc:
                expected_pairs += doc.page_count // 2
        except Exception:
            pass

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(uploaded_files)}</strong><span>PDF upload(s)</span></div>
            <div class="metric-box"><strong>{expected_pairs}</strong><span>two-page output file(s)</span></div>
            <div class="metric-box"><strong>ZIP</strong><span>renamed download</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.caption("After upload, click the button below. The app will create one renamed PDF for each two-page pair.")

    if st.button("Split and Rename PDFs", type="primary", use_container_width=True):
        progress = st.progress(0, text="Splitting PDFs...")
        try:
            zip_bytes, result_rows, logs = split_and_rename_pdfs(uploaded_files, include_odd_final_page, use_fallback_names)
            st.session_state["beam_zip_bytes"] = zip_bytes
            st.session_state["beam_result_rows"] = result_rows
            st.session_state["beam_result_logs"] = logs
            progress.progress(100, text="Finished")
        except Exception as exc:
            st.session_state["beam_zip_bytes"] = None
            st.session_state["beam_result_rows"] = []
            st.session_state["beam_result_logs"] = [f"Processing failed: {exc}"]
            progress.progress(100, text="Failed")

    result_rows = st.session_state.get("beam_result_rows", [])
    logs = st.session_state.get("beam_result_logs", [])

    if logs:
        with st.expander("Processing log", expanded=True):
            for log in logs:
                st.write(log)
        if not st.session_state.get("beam_zip_bytes"):
            st.error("No download was created. Check the processing log above.")

    if result_rows:
        st.subheader("Created Files")
        st.dataframe(result_rows, use_container_width=True, hide_index=True)

    if st.session_state.get("beam_zip_bytes"):
        st.success("Your split and renamed PDFs are ready.")
        st.download_button(
            "Download Beam Split Results",
            data=st.session_state["beam_zip_bytes"],
            file_name="beam_split_results.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )


def render_affidavit_pdf_splitter() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Affidavit PDF Splitter</h1>
            <p>Upload affidavit PDFs, split them into two-page files, and name each output from the OFN number at the bottom of the second page.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("Processing rules", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            include_odd_final_page = st.checkbox(
                "Include odd final page as a single-page PDF",
                value=False,
                key="affidavit_include_odd_final_page",
            )
        with col2:
            use_fallback_names = st.checkbox(
                "Use fallback name if OFN number is not found",
                value=False,
                key="affidavit_use_fallback_names",
            )

    st.markdown('<div class="section-label">Upload PDFs</div>', unsafe_allow_html=True)
    upload_tabs = st.tabs(["Single or multiple PDF files", "Folder of PDFs"])
    with upload_tabs[0]:
        file_uploads = st.file_uploader(
            "Choose one PDF or several PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            help="Use this when you have one PDF or a few PDFs selected manually.",
            key="affidavit_pdf_file_uploads",
        )
    with upload_tabs[1]:
        folder_uploads = st.file_uploader(
            "Choose a folder",
            type=["pdf"],
            accept_multiple_files="directory",
            help="Use this when you want to upload all PDFs from a folder.",
            key="affidavit_pdf_folder_uploads",
        )

    uploaded_files = list(file_uploads or []) + list(folder_uploads or [])
    upload_signature = uploaded_files_signature(uploaded_files)
    if st.session_state.get("affidavit_upload_signature") != upload_signature:
        st.session_state["affidavit_upload_signature"] = upload_signature
        st.session_state["affidavit_zip_bytes"] = None
        st.session_state["affidavit_result_rows"] = []
        st.session_state["affidavit_result_logs"] = []

    if not uploaded_files:
        left, right = st.columns(2)
        with left:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>What it does</h3>
                    <p>Splits page pairs like 1-2, 3-4, 5-6 and names each output from the OFN number plus W at the end.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with right:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>Output names</h3>
                    <p>If the second page says OFN: 12942, the downloaded file will be named 12942W.pdf.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        return

    expected_pairs = 0
    for uploaded_file in uploaded_files:
        try:
            with fitz.open(stream=uploaded_file.getvalue(), filetype="pdf") as doc:
                expected_pairs += doc.page_count // 2
        except Exception:
            pass

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(uploaded_files)}</strong><span>PDF upload(s)</span></div>
            <div class="metric-box"><strong>{expected_pairs}</strong><span>two-page output file(s)</span></div>
            <div class="metric-box"><strong>OFN + W</strong><span>filename format</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.caption("After upload, click the button below. The app will create one renamed PDF for each two-page pair.")

    if st.button("Split Affidavit PDFs", type="primary", use_container_width=True):
        progress = st.progress(0, text="Splitting PDFs...")
        try:
            zip_bytes, result_rows, logs = split_and_rename_ofn_pdfs(uploaded_files, include_odd_final_page, use_fallback_names)
            st.session_state["affidavit_zip_bytes"] = zip_bytes
            st.session_state["affidavit_result_rows"] = result_rows
            st.session_state["affidavit_result_logs"] = logs
            progress.progress(100, text="Finished")
        except Exception as exc:
            st.session_state["affidavit_zip_bytes"] = None
            st.session_state["affidavit_result_rows"] = []
            st.session_state["affidavit_result_logs"] = [f"Processing failed: {exc}"]
            progress.progress(100, text="Failed")

    result_rows = st.session_state.get("affidavit_result_rows", [])
    logs = st.session_state.get("affidavit_result_logs", [])

    if logs:
        with st.expander("Processing log", expanded=True):
            for log in logs:
                st.write(log)
        if not st.session_state.get("affidavit_zip_bytes"):
            st.error("No download was created. Check the processing log above.")

    if result_rows:
        st.subheader("Created Files")
        st.dataframe(result_rows, use_container_width=True, hide_index=True)

    if st.session_state.get("affidavit_zip_bytes"):
        st.success("Your affidavit PDFs are split and renamed.")
        st.download_button(
            "Download Affidavit Split Results",
            data=st.session_state["affidavit_zip_bytes"],
            file_name="affidavit_split_results.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )


def render_home(tools: list[ToolDefinition]) -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Resolution Law Tools</h1>
            <p>A clean web toolbox for PDF, document, and office workflows. Use the PDF splitters and spreadsheet tools as the team needs them.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("Available Tools")
    columns = st.columns(2)
    for index, tool in enumerate(tools):
        with columns[index % 2]:
            st.markdown(
                f"""
                <div class="tool-card clickable">
                    <h3>{tool.name}</h3>
                    <p>{tool.description}</p>
                    <p class="small-muted">{tool.category} tool</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.button(tool.name, key=f"open_{tool.tool_id}", use_container_width=True):
                st.session_state["active_tool_label"] = tool.name
                st.rerun()

    st.info("Future tools can be added by adding another render function and ToolDefinition in app.py.")


def render_future_tools_guide() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>Add Future Tools</h1>
            <p>Resolution Law Tools is set up as a toolbox. Add another workflow by creating a render function and registering it in the tools list.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.write("Add a new function like this:")
    st.code(
        '''def render_new_tool() -> None:
    st.markdown("""
    <div class="hub-hero">
        <h1>New Tool</h1>
        <p>Short description of what this tool does.</p>
    </div>
    """, unsafe_allow_html=True)
    st.write("Build the workflow here.")
''',
        language="python",
    )

    st.write("Then add a registry entry:")
    st.code(
        '''ToolDefinition(
    tool_id="new-tool",
    name="New Tool",
    category="Documents",
    description="Short plain-English description.",
    render=render_new_tool,
)''',
        language="python",
    )


# ============================================================================
# File Name Lister — Resolution Law Tools
# Takes one or more uploaded files (or a whole folder of files) and builds a
# spreadsheet (.xlsx) listing every file name. No file contents are read; this
# just inventories names, types, and sizes so the team can paste the list into
# a tracker or use it as a checklist.
# ============================================================================
def _human_size(num_bytes: int) -> str:
    size = float(num_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024 or unit == "TB":
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{num_bytes} B"


def build_file_list_rows(uploaded_files) -> list[dict]:
    rows: list[dict] = []
    for index, uploaded_file in enumerate(uploaded_files, start=1):
        raw_name = uploaded_file.name or ""
        # Folder uploads can include a relative path like "subfolder/file.pdf".
        normalized = raw_name.replace("\\", "/")
        folder = ""
        base_name = normalized
        if "/" in normalized:
            folder, base_name = normalized.rsplit("/", 1)
        stem, dot, ext = base_name.rpartition(".")
        if not dot:
            stem, ext = base_name, ""
        try:
            size_bytes = len(uploaded_file.getvalue())
        except Exception:
            size_bytes = 0
        rows.append(
            {
                "#": index,
                "File Name": base_name,
                "Name Without Extension": stem,
                "Extension": ext.lower(),
                "Folder": folder,
                "Size": _human_size(size_bytes),
                "Size (Bytes)": size_bytes,
            }
        )
    return rows


def build_file_list_excel(rows: list[dict]) -> bytes:
    headers = [
        "#",
        "File Name",
        "Name Without Extension",
        "Extension",
        "Folder",
        "Size",
        "Size (Bytes)",
    ]
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Names"
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = _Font(bold=True, color="FFFFFF")
        c.fill = _PatternFill("solid", fgColor="1565C0")
        c.alignment = _Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for ci, h in enumerate(headers, 1):
        samples = [len(str(h))] + [min(len(str(r.get(h, ""))), 80) for r in rows[:500]]
        ml = max(samples) if samples else 12
        ws.column_dimensions[_openpyxl.utils.get_column_letter(ci)].width = min(max(10, ml + 2), 80)
    ws.freeze_panes = "A2"
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_file_name_lister() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>File Name Lister</h1>
            <p>Upload one file, several files, or a whole folder, and download a spreadsheet that lists every file name.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("How it works", expanded=False):
        st.markdown(
            "- Upload a single file, multiple files, or an entire folder.\n"
            "- The tool reads only the file names, types, and sizes — never the contents.\n"
            "- You get an .xlsx with one row per file: name, name without extension, extension, folder, and size.\n"
            "- Use it to build an index, a checklist, or to paste names into another tracker."
        )

    st.markdown('<div class="section-label">Upload Files</div>', unsafe_allow_html=True)
    upload_tabs = st.tabs(["Single or multiple files", "Folder of files"])
    with upload_tabs[0]:
        file_uploads = st.file_uploader(
            "Choose one file or several files",
            accept_multiple_files=True,
            help="Any file type is allowed. Only the names are recorded.",
            key="filelist_file_uploads",
        )
    with upload_tabs[1]:
        folder_uploads = st.file_uploader(
            "Choose a folder",
            accept_multiple_files="directory",
            help="Use this to upload every file from a folder at once.",
            key="filelist_folder_uploads",
        )

    uploaded_files = list(file_uploads or []) + list(folder_uploads or [])
    upload_signature = uploaded_files_signature(uploaded_files)
    if st.session_state.get("filelist_upload_signature") != upload_signature:
        st.session_state["filelist_upload_signature"] = upload_signature
        st.session_state["filelist_excel_bytes"] = None
        st.session_state["filelist_rows"] = []

    if not uploaded_files:
        left, right = st.columns(2)
        with left:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>What it does</h3>
                    <p>Turns a pile of files into a clean spreadsheet listing every file name, type, and size.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with right:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>Names only</h3>
                    <p>The tool never opens or reads the files. It only records their names and sizes.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        return

    rows = build_file_list_rows(uploaded_files)
    st.session_state["filelist_rows"] = rows

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(rows)}</strong><span>file(s) found</span></div>
            <div class="metric-box"><strong>{len({r['Extension'] for r in rows})}</strong><span>file type(s)</span></div>
            <div class="metric-box"><strong>XLSX</strong><span>spreadsheet download</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("File Names")
    st.dataframe(rows, use_container_width=True, hide_index=True)

    if st.button("Create Spreadsheet", type="primary", use_container_width=True):
        try:
            st.session_state["filelist_excel_bytes"] = build_file_list_excel(rows)
        except Exception as exc:
            st.session_state["filelist_excel_bytes"] = None
            st.error(f"Could not build the spreadsheet: {exc}")

    if st.session_state.get("filelist_excel_bytes"):
        st.success("Your file name spreadsheet is ready.")
        st.download_button(
            "Download File Name Spreadsheet",
            data=st.session_state["filelist_excel_bytes"],
            file_name="file_names.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


# ============================================================================
# Account Document Builder — Resolution Law Tools
# Upload a batch of files (or a whole "Doc Builder" folder). The tool reads the
# account number from each file (from an "Acct#####" token in the file name or
# from an account-number folder such as 16638/Images/...), groups every file by
# account, then merges each account's files into ONE combined PDF named
# "<account>final.pdf" (complaint letter first, then supporting documents).
# Excel and image files are converted to PDF pages before merging.
# ============================================================================
ACCT_IN_NAME = re.compile(r"Acct[\s_\-]*([0-9]{3,7})", re.IGNORECASE)
ACCT_FOLDER_SEGMENT = re.compile(r"^[0-9]{3,7}$")
LETTER_HINT = re.compile(r"Ltr[\s_\-]*[0-9]+", re.IGNORECASE)
DOCBUILD_PDF_EXTS = {"pdf"}
DOCBUILD_EXCEL_EXTS = {"xls", "xlsx", "xlsm"}
DOCBUILD_IMAGE_EXTS = {"png", "jpg", "jpeg", "tif", "tiff", "bmp", "gif"}
DOCBUILD_HTML_EXTS = {"html", "htm"}
DOCBUILD_WORD_EXTS = {"doc", "docx"}
DOCBUILD_ARCHIVE_EXTS = {"zip"}


def _split_relative_name(raw_name: str) -> tuple[list[str], str]:
    """Return (folder_segments, base_name) from an upload name that may carry a
    relative path like 'DocBuilder/16638/Images/file.pdf'."""
    normalized = (raw_name or "").replace("\\", "/")
    parts = [p for p in normalized.split("/") if p]
    if not parts:
        return [], ""
    return parts[:-1], parts[-1]


def _file_extension(base_name: str) -> str:
    return base_name.rsplit(".", 1)[-1].lower() if "." in base_name else ""


def extract_account_number(raw_name: str) -> str | None:
    """Find the account number for a file: first from an 'Acct#####' token in the
    file name, otherwise from the nearest all-digits folder segment."""
    folders, base = _split_relative_name(raw_name)
    match = ACCT_IN_NAME.search(base)
    if match:
        return match.group(1)
    for segment in reversed(folders):
        if ACCT_FOLDER_SEGMENT.match(segment):
            return segment
    return None


def is_letter_file(base_name: str) -> bool:
    """A complaint letter carries both an Acct token and an Ltr token."""
    return bool(ACCT_IN_NAME.search(base_name) and LETTER_HINT.search(base_name))


def _file_sort_key(base_name: str) -> tuple[int, str]:
    """Letters first (0), then everything else (1); supporting documents start
    with a date/timestamp so a plain name sort keeps them chronological."""
    return (0 if is_letter_file(base_name) else 1, base_name.lower())


def _safe_pdf_text(value: str) -> str:
    """PyMuPDF base-14 fonts are Latin-1; replace anything outside it."""
    return value.encode("latin-1", "replace").decode("latin-1")


def _xls_cell_to_str(value) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return "" if value is None else str(value)


def _read_spreadsheet_rows(data: bytes, ext: str, max_rows: int = 3000) -> list[tuple[str, list[list[str]]]]:
    """Return [(sheet_title, rows-of-strings), ...] for a spreadsheet, or []."""
    sheets: list[tuple[str, list[list[str]]]] = []
    if ext in {"xlsx", "xlsm"}:
        wb = _openpyxl.load_workbook(_io.BytesIO(data), data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for record in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in record])
                if len(rows) >= max_rows:
                    break
            sheets.append((ws.title, rows))
    elif ext == "xls":
        import xlrd  # type: ignore

        book = xlrd.open_workbook(file_contents=data)
        for sheet in book.sheets():
            rows = []
            for r in range(min(sheet.nrows, max_rows)):
                rows.append([_xls_cell_to_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
            sheets.append((sheet.name, rows))
    return sheets


_LIBREOFFICE_BIN_CACHE: list = []


def _find_libreoffice_binary() -> str | None:
    """Locate a LibreOffice/soffice executable, if one is installed."""
    if _LIBREOFFICE_BIN_CACHE:
        return _LIBREOFFICE_BIN_CACHE[0] or None
    found = None
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            found = path
            break
    _LIBREOFFICE_BIN_CACHE.append(found or "")
    return found


def convert_office_to_pdf_via_libreoffice(data: bytes, base_name: str) -> bytes | None:
    """Convert any office file (xls, xlsx, doc, docx, html, ...) to a properly
    formatted PDF using a headless LibreOffice, if it is available. Returns None
    if LibreOffice is not installed or the conversion fails. This is the highest
    fidelity path; pure-Python fallbacks below run when it is unavailable."""
    soffice = _find_libreoffice_binary()
    if not soffice:
        return None
    ext = _file_extension(base_name) or "bin"
    try:
        with tempfile.TemporaryDirectory() as tmp:
            src_path = os.path.join(tmp, f"input.{ext}")
            with open(src_path, "wb") as fh:
                fh.write(data)
            profile = "file://" + os.path.join(tmp, "lo_profile")
            subprocess.run(
                [
                    soffice, "--headless", "--norestore", "--nolockcheck", "--nodefault",
                    f"-env:UserInstallation={profile}",
                    "--convert-to", "pdf", "--outdir", tmp, src_path,
                ],
                capture_output=True, timeout=180,
            )
            for fname in os.listdir(tmp):
                if fname.lower().endswith(".pdf"):
                    with open(os.path.join(tmp, fname), "rb") as fh:
                        out = fh.read()
                    if out:
                        return out
            return None
    except Exception:
        return None


def _render_rows_to_pdf_bytes(sheets, base_name: str) -> bytes | None:
    """Fallback renderer: draw spreadsheet rows as monospaced text pages."""
    if not sheets:
        return None
    try:
        doc = fitz.open()
        page_w, page_h = 792.0, 612.0  # landscape Letter
        margin = 36.0
        line_h = 11.0
        max_chars = 170
        for title, rows in sheets:
            page = doc.new_page(width=page_w, height=page_h)
            y = margin
            page.insert_text((margin, y), _safe_pdf_text(f"{base_name} - {title}"), fontsize=10, fontname="hebo")
            y += line_h * 2
            if not rows:
                page.insert_text((margin, y), "(empty sheet)", fontsize=8, fontname="cour")
                continue
            for row in rows:
                line = " | ".join(row)
                if len(line) > max_chars:
                    line = line[: max_chars - 3] + "..."
                if y > page_h - margin:
                    page = doc.new_page(width=page_w, height=page_h)
                    y = margin
                page.insert_text((margin, y), _safe_pdf_text(line), fontsize=8, fontname="cour")
                y += line_h
        out = doc.tobytes()
        doc.close()
        return out
    except Exception:
        return None


def _html_string_to_pdf_bytes(html: str, landscape: bool = False) -> bytes | None:
    """Render an HTML string to a paginated PDF using PyMuPDF's Story engine.
    The Story engine wraps long cell text and paginates automatically, so no
    data is ever cut off the page."""
    if not html:
        html = "<p>(empty document)</p>"
    try:
        story = fitz.Story(html=html)
        buf = BytesIO()
        writer = fitz.DocumentWriter(buf)
        mediabox = fitz.paper_rect("letter-l" if landscape else "letter")
        where = mediabox + (36, 36, -36, -36)
        more = 1
        while more:
            device = writer.begin_page(mediabox)
            more, _ = story.place(where)
            story.draw(device)
            writer.end_page()
        writer.close()
        return buf.getvalue()
    except Exception:
        return None


def _spreadsheet_to_html_table(sheets, base_name: str, max_cols_per_block: int = 7) -> str:
    """Build styled, bordered HTML tables from spreadsheet rows. Every column and
    row is included. Sheets wider than max_cols_per_block are split into column
    blocks (the first column repeats in each block for context) so that even very
    wide spreadsheets fit the page and no column is clipped or cut off."""
    import html as _html

    css = (
        "<style>"
        "h3{font-family:Helvetica,Arial,sans-serif;font-size:11px;margin:8px 0 4px 0;}"
        "h4{font-family:Helvetica,Arial,sans-serif;font-size:8.5px;color:#444;margin:8px 0 3px 0;}"
        "table{border-collapse:collapse;width:100%;font-family:Helvetica,Arial,sans-serif;font-size:8px;}"
        "th,td{border:1px solid #9aa0a6;padding:3px 4px;text-align:left;vertical-align:top;}"
        "th{background-color:#1565C0;color:#ffffff;font-weight:bold;}"
        "tr:nth-child(even) td{background-color:#f2f5fa;}"
        "</style>"
    )

    def esc(v):
        return _html.escape(str(v))

    def render_block(header, body, col_indices):
        out = ["<table>"]
        out.append("<tr>" + "".join(f"<th>{esc(header[c])}</th>" for c in col_indices) + "</tr>")
        for row in body:
            out.append("<tr>" + "".join(f"<td>{esc(row[c])}</td>" for c in col_indices) + "</tr>")
        out.append("</table>")
        return "".join(out)

    parts = [css]
    for title, rows in sheets:
        parts.append(f"<h3>{esc(base_name)} &mdash; {esc(title)}</h3>")
        if not rows:
            parts.append("<p>(empty sheet)</p>")
            continue
        ncols = max((len(r) for r in rows), default=0)
        if ncols == 0:
            parts.append("<p>(empty sheet)</p>")
            continue
        padded = [list(r) + [""] * (ncols - len(r)) for r in rows]
        header, body = padded[0], padded[1:]

        if ncols <= max_cols_per_block:
            parts.append(render_block(header, body, list(range(ncols))))
            continue

        # Wide sheet: split into column blocks, repeating column 0 for context.
        rest = list(range(1, ncols))
        per_block = max(1, max_cols_per_block - 1)
        blocks = [rest[i:i + per_block] for i in range(0, len(rest), per_block)]
        for n, block in enumerate(blocks, start=1):
            col_indices = [0] + block
            first, last = block[0] + 1, block[-1] + 1
            parts.append(
                f"<h4>Columns 1 and {first}-{last} of {ncols} (block {n} of {len(blocks)})</h4>"
            )
            parts.append(render_block(header, body, col_indices))
    return "".join(parts)


REDACTION_TEXT = "[REDACTED]"
# Formatted SSNs anywhere in a cell (123-45-6789 or 123 45 6789).
_SSN_FORMATTED = re.compile(r"\b\d{3}[-\s]\d{2}[-\s]\d{4}\b")
# Column-header cues for whole-column redaction.
_SSN_HEADER = re.compile(r"(ssn|social\s*sec|soc\s*sec|\btin\b|tax\s*id)", re.I)
_DOB_HEADER = re.compile(r"(date\s*of\s*birth|d\.?o\.?b\.?|birth\s*date|birthday|\bdob\b|\bbirth\b)", re.I)


def _redact_sheets(sheets):
    """Redact SSNs and dates of birth from spreadsheet rows before rendering.

    - Any column whose HEADER looks like SSN or Date-of-Birth has every value
      replaced with [REDACTED].
    - Any cell anywhere containing a formatted SSN (123-45-6789 / 123 45 6789)
      has that value replaced, even if the column is not labeled.
    DOB is column-header driven on purpose, so ordinary dates (payment/posted
    dates, etc.) are NOT touched. The redacted text never reaches the PDF."""
    out = []
    for title, rows in sheets:
        if not rows:
            out.append((title, rows))
            continue
        header = rows[0]
        redact_cols = set()
        for i, h in enumerate(header):
            hs = str(h)
            if _SSN_HEADER.search(hs) or _DOB_HEADER.search(hs):
                redact_cols.add(i)
        new_rows = [header]
        for row in rows[1:]:
            nr = []
            for i, cell in enumerate(row):
                cv = "" if cell is None else str(cell)
                if i in redact_cols and cv.strip():
                    cv = REDACTION_TEXT
                elif cv:
                    cv = _SSN_FORMATTED.sub(REDACTION_TEXT, cv)
                nr.append(cv)
            new_rows.append(nr)
        out.append((title, new_rows))
    return out


def convert_spreadsheet_to_pdf_bytes(data: bytes, base_name: str) -> bytes | None:
    """Spreadsheet -> PDF rendered as a nicely formatted, bordered table
    (landscape, wrapping, auto-paginated) so every value is included and nothing
    is cut off. SSNs and dates of birth are redacted from the data before the PDF
    is built. (The spreadsheet always goes through this renderer, never an
    external converter, so redaction is guaranteed.)"""
    ext = _file_extension(base_name)
    try:
        sheets = _read_spreadsheet_rows(data, ext)
    except Exception:
        return None
    if not sheets:
        return None
    sheets = _redact_sheets(sheets)
    table_pdf = _html_string_to_pdf_bytes(_spreadsheet_to_html_table(sheets, base_name), landscape=True)
    if table_pdf is not None:
        return table_pdf
    return _render_rows_to_pdf_bytes(sheets, base_name)  # last-resort monospaced text


def _make_placeholder_pdf_bytes(title: str, note: str) -> bytes | None:
    """Create a single-page PDF noting a file that could not be merged, so it is
    still represented (in order) inside the combined account PDF."""
    try:
        doc = fitz.open()
        page = doc.new_page(width=612, height=792)
        page.insert_text((72, 96), _safe_pdf_text(title), fontsize=12, fontname="hebo")
        wrapped = _safe_pdf_text(note)
        rect = fitz.Rect(72, 120, 540, 300)
        try:
            page.insert_textbox(rect, wrapped, fontsize=10, fontname="helv")
        except Exception:
            page.insert_text((72, 130), wrapped[:120], fontsize=10, fontname="helv")
        out = doc.tobytes()
        doc.close()
        return out
    except Exception:
        return None


def convert_html_to_pdf_bytes(data: bytes, base_name: str) -> bytes | None:
    """HTML -> PDF. LibreOffice first; else PyMuPDF's HTML Story engine."""
    via_office = convert_office_to_pdf_via_libreoffice(data, base_name)
    if via_office is not None:
        return via_office
    try:
        html = data.decode("utf-8")
    except Exception:
        html = data.decode("latin-1", "replace")
    return _html_string_to_pdf_bytes(html)


def convert_word_to_pdf_bytes(data: bytes, base_name: str) -> bytes | None:
    """Word -> PDF. LibreOffice first (handles .doc and .docx); else .docx is
    converted via mammoth (docx -> HTML) and rendered with PyMuPDF."""
    via_office = convert_office_to_pdf_via_libreoffice(data, base_name)
    if via_office is not None:
        return via_office
    if _file_extension(base_name) == "docx":
        try:
            import mammoth  # type: ignore

            html = mammoth.convert_to_html(BytesIO(data)).value
            return _html_string_to_pdf_bytes(html or "")
        except Exception:
            return None
    return None  # legacy .doc has no pure-Python path


def convert_image_to_pdf_bytes(data: bytes) -> bytes | None:
    """Image -> PDF. Pillow first (handles multi-page TIFF); else PyMuPDF."""
    try:
        from PIL import Image  # type: ignore

        img = Image.open(BytesIO(data))
        frames = getattr(img, "n_frames", 1)
        pages = []
        for index in range(frames):
            try:
                img.seek(index)
            except Exception:
                break
            pages.append(img.convert("RGB"))
        if pages:
            buf = BytesIO()
            pages[0].save(buf, format="PDF", save_all=True, append_images=pages[1:])
            out = buf.getvalue()
            if out:
                return out
    except Exception:
        pass
    try:
        img_doc = fitz.open(stream=data, filetype=None)
        try:
            pdf_bytes = img_doc.convert_to_pdf()
        finally:
            img_doc.close()
        return pdf_bytes
    except Exception:
        return None


def convert_to_pdf_bytes(data: bytes, base_name: str) -> tuple[bytes | None, str]:
    """Dispatch a single file to the right converter.
    Returns (pdf_bytes_or_None, status) where status is one of:
    'pdf' (already a PDF), 'converted', 'unsupported', 'failed'."""
    ext = _file_extension(base_name)
    if ext in DOCBUILD_PDF_EXTS:
        return data, "pdf"
    if ext in DOCBUILD_EXCEL_EXTS:
        out = convert_spreadsheet_to_pdf_bytes(data, base_name)
        return (out, "converted") if out else (None, "failed")
    if ext in DOCBUILD_HTML_EXTS:
        out = convert_html_to_pdf_bytes(data, base_name)
        return (out, "converted") if out else (None, "failed")
    if ext in DOCBUILD_WORD_EXTS:
        out = convert_word_to_pdf_bytes(data, base_name)
        return (out, "converted") if out else (None, "failed")
    if ext in DOCBUILD_IMAGE_EXTS:
        out = convert_image_to_pdf_bytes(data)
        return (out, "converted") if out else (None, "failed")
    return None, "unsupported"


def _expand_uploads(base_name: str, data: bytes) -> list[tuple[str, bytes]]:
    """Expand archives into their contained files. A .zip yields one entry per
    file inside it; everything else yields itself unchanged."""
    if _file_extension(base_name) not in DOCBUILD_ARCHIVE_EXTS:
        return [(base_name, data)]
    try:
        archive = zipfile.ZipFile(BytesIO(data))
    except Exception:
        return [(base_name, data)]
    entries: list[tuple[str, bytes]] = []
    for info in archive.infolist():
        if info.is_dir():
            continue
        entry = info.filename.replace("\\", "/").split("/")[-1]
        if not entry:
            continue
        try:
            entries.append((f"{base_name} > {entry}", archive.read(info)))
        except Exception:
            continue
    return entries or [(base_name, data)]


@dataclass
class AccountBuildResult:
    account: str
    files_found: int
    pdfs_merged: int
    converted: int
    skipped: int
    pages: int
    output_file: str
    status: str


def _longest_common_folder_prefix(seglists: list[list[str]]) -> list[str]:
    """Longest shared leading folder segments across all uploads (used to strip a
    wrapping upload folder so the real sub-folders are detected correctly)."""
    seglists = [s for s in seglists]
    if not seglists:
        return []
    prefix = list(seglists[0])
    for segs in seglists[1:]:
        i = 0
        while i < len(prefix) and i < len(segs) and prefix[i] == segs[i]:
            i += 1
        prefix = prefix[:i]
        if not prefix:
            break
    return prefix


def _extract_group_number(name: str) -> str | None:
    """Pull an account-style number from a file name: an 'Acct####' token first,
    otherwise the longest standalone run of >=3 digits."""
    m = ACCT_IN_NAME.search(name)
    if m:
        return m.group(1)
    runs = re.findall(r"(?<!\d)(\d{3,})(?!\d)", name)
    if runs:
        return max(runs, key=len)
    return None


def _match_folder_for_loose(base_name: str, folder_names) -> str | None:
    """Find the sub-folder whose name appears as a standalone token in an outside
    file's name (e.g. folder '16638' matches 'Letter..._Acct16638_Ltr.pdf')."""
    matches = []
    for fn in folder_names:
        if fn.isdigit():
            if re.search(r"(?<!\d)" + re.escape(fn) + r"(?!\d)", base_name):
                matches.append(fn)
        elif re.search(r"(?<![A-Za-z0-9])" + re.escape(fn) + r"(?![A-Za-z0-9])", base_name, re.I):
            matches.append(fn)
    if not matches:
        return None
    matches.sort(key=len, reverse=True)  # most specific (longest) wins
    return matches[0]


def _folder_account_key(folder_name: str) -> str:
    """The group key for a sub-folder: its account number if the folder name
    contains one (e.g. '16638', '16638 - John Doe', 'Acct16638' all -> '16638'),
    otherwise the folder name itself."""
    return _extract_group_number(folder_name) or folder_name


def _group_uploads(uploaded_files):
    """Shared grouping used by both the build and the on-screen preview so their
    numbers always agree. Files are grouped by the account number of the sub-folder
    they live in; outside files are matched to a group by the account number in
    their name. Returns (inside, front, group_names):
      inside[group] -> list of (base_name, upload) for files inside that folder
      front[group]  -> list of (base_name, upload) for OUTSIDE files matched to it
      group_names   -> every group that will become a <group>final.pdf
    """
    seglists = [_split_relative_name(uf.name or "")[0] for uf in uploaded_files]
    root_len = len(_longest_common_folder_prefix(seglists))

    inside: dict[str, list[tuple[str, object]]] = {}
    loose: list[tuple[str, object]] = []
    for uploaded_file in uploaded_files:
        folders, base = _split_relative_name(uploaded_file.name or "")
        stripped = folders[root_len:]
        if stripped:
            key = _folder_account_key(stripped[0])
            inside.setdefault(key, []).append((base, uploaded_file))
        else:
            loose.append((base, uploaded_file))

    account_keys = set(inside.keys())
    front: dict[str, list[tuple[str, object]]] = {}
    group_names = set(account_keys)
    for base, uploaded_file in loose:
        num = _extract_group_number(base)
        group = None
        if num and num in account_keys:
            group = num
        if group is None:
            # fall back: any numeric account key that appears as a token in the name
            for k in account_keys:
                if k.isdigit() and re.search(r"(?<!\d)" + re.escape(k) + r"(?!\d)", base):
                    group = k
                    break
        if group is None:
            group = num or sanitize_filename_stem(base, "document")
        front.setdefault(group, []).append((base, uploaded_file))
        group_names.add(group)
    return inside, front, group_names


def build_account_final_pdfs(uploaded_files) -> tuple[bytes | None, list[dict], list[str], list[str]]:
    """Build one combined PDF per sub-folder.

    Grouping is folder-driven: every file INSIDE a sub-folder is combined for that
    folder, and any file OUTSIDE the sub-folders is prepended to the FRONT of the
    matching folder's PDF (matched by the folder's number appearing in the outside
    file's name). Output files are named <folder>final.pdf and placed flat in the
    ZIP. Returns (zip_bytes_or_None, result_rows, processing_log, unmatched_names)."""
    # Folder-driven grouping (shared with the preview so counts always agree).
    inside, front, group_names = _group_uploads(uploaded_files)

    results: list[dict] = []
    log: list[str] = []
    unmatched: list[str] = []
    written = 0
    output = BytesIO()

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for group in sorted(group_names):
            front_items = sorted(front.get(group, []), key=lambda pair: pair[0].lower())
            body_items = sorted(inside.get(group, []), key=lambda pair: pair[0].lower())
            items = front_items + body_items  # outside files first, then folder files

            merged = fitz.open()
            pdfs_merged = 0
            converted = 0
            skipped = 0
            placeholders = 0

            # Read every file, expanding any .zip archives, preserving order.
            flat: list[tuple[str, bytes]] = []
            for base, uploaded_file in items:
                try:
                    data = uploaded_file.getvalue()
                except Exception as exc:
                    skipped += 1
                    log.append(f"{group}: skipped {base} (could not read upload: {exc})")
                    continue
                flat.extend(_expand_uploads(base, data))

            def _add_placeholder(display_name: str, note: str) -> None:
                """Insert a labeled placeholder page so a file is never dropped."""
                nonlocal placeholders, skipped
                ph = _make_placeholder_pdf_bytes(display_name, note)
                if ph is None:
                    skipped += 1
                    log.append(f"{group}: skipped {display_name} ({note})")
                    return
                try:
                    ph_doc = fitz.open(stream=ph, filetype="pdf")
                    try:
                        merged.insert_pdf(ph_doc)
                    finally:
                        ph_doc.close()
                    placeholders += 1
                    log.append(f"{group}: placeholder page added for {display_name} ({note})")
                except Exception:
                    skipped += 1
                    log.append(f"{group}: skipped {display_name} ({note})")

            for name, data in flat:
                effective = name.split(" > ")[-1]
                pdf_bytes, status = convert_to_pdf_bytes(data, effective)
                if status == "converted":
                    converted += 1
                elif status == "unsupported":
                    _add_placeholder(name, f"Unsupported file type '.{_file_extension(effective)}' - included as a placeholder")
                    continue
                elif status == "failed":
                    _add_placeholder(name, "Could not convert to PDF - included as a placeholder")
                    continue

                if pdf_bytes is None:
                    _add_placeholder(name, "No content could be read - included as a placeholder")
                    continue

                try:
                    src = fitz.open(stream=pdf_bytes, filetype="pdf")
                    try:
                        if src.page_count == 0:
                            _add_placeholder(name, "Source PDF is empty or corrupt (0 pages) - included as a placeholder")
                        else:
                            merged.insert_pdf(src)
                            pdfs_merged += 1
                    finally:
                        src.close()
                except Exception:
                    _add_placeholder(name, "Source file is empty or corrupt - included as a placeholder")

            page_count = merged.page_count
            if page_count > 0:
                out_name = f"{group}final.pdf"
                archive.writestr(out_name, merged.tobytes())
                written += 1
                status = "OK"
            else:
                out_name = "(none)"
                status = "No PDF pages produced"
            merged.close()

            results.append(
                {
                    "Account": group,
                    "Outside (front)": len(front.get(group, [])),
                    "In folder": len(inside.get(group, [])),
                    "PDFs Merged": pdfs_merged,
                    "Converted": converted,
                    "Placeholders": placeholders,
                    "Skipped": skipped,
                    "Pages": page_count,
                    "Output File": out_name,
                    "Status": status,
                }
            )

    if written == 0:
        return None, results, log, unmatched
    return output.getvalue(), results, log, unmatched


def render_file_combiner() -> None:
    st.markdown(
        """
        <div class="hub-hero">
            <h1>File Combiner</h1>
            <p>Upload your files or whole Doc Builder folder. The tool sorts every file by account number and merges each account into one combined PDF named &lt;account&gt;final.pdf.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander("How it works", expanded=True):
        st.markdown(
            "- Upload individual files, several files, or an entire account folder.\n"
            "- The account number is read from an **Acct#####** token in the file name "
            "(such as `..._Acct16638_Ltr31648.pdf`) or from an **account-number folder** "
            "(such as `16638/Images/...`).\n"
            "- Every file is sorted into its account, then merged into one PDF named "
            "**`<account>final.pdf`** - the complaint letter first, then supporting "
            "documents in date order.\n"
            "- PDFs are merged directly. Excel and image files are converted to PDF pages "
            "first. Anything that can't be converted is listed in the processing log.\n"
            "- You get a single ZIP containing all the combined PDFs together in one "
            "folder, each named `<account>final.pdf`."
        )

    st.markdown('<div class="section-label">Upload Files</div>', unsafe_allow_html=True)
    upload_tabs = st.tabs(["Single or multiple files", "Folder of files"])
    with upload_tabs[0]:
        file_uploads = st.file_uploader(
            "Choose one file or several files",
            accept_multiple_files=True,
            help="Letters and supporting documents. Account numbers are read from names like 'Acct16638'.",
            key="docbuild_file_uploads",
        )
    with upload_tabs[1]:
        folder_uploads = st.file_uploader(
            "Choose a folder",
            accept_multiple_files="directory",
            help="Upload a whole Doc Builder folder; account-number subfolders are detected automatically.",
            key="docbuild_folder_uploads",
        )

    uploaded_files = list(file_uploads or []) + list(folder_uploads or [])
    upload_signature = uploaded_files_signature(uploaded_files)
    if st.session_state.get("docbuild_upload_signature") != upload_signature:
        st.session_state["docbuild_upload_signature"] = upload_signature
        st.session_state["docbuild_zip_bytes"] = None
        st.session_state["docbuild_results"] = []
        st.session_state["docbuild_log"] = []
        st.session_state["docbuild_unmatched"] = []

    if not uploaded_files:
        left, right = st.columns(2)
        with left:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>What it does</h3>
                    <p>Sorts a pile of files by account number and builds one combined PDF per account, named &lt;account&gt;final.pdf.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with right:
            st.markdown(
                """
                <div class="tool-card">
                    <h3>Letter first</h3>
                    <p>Each account's complaint letter leads, followed by its supporting documents in date order. Excel and images are converted to PDF.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        return

    _inside, _front, _group_names = _group_uploads(uploaded_files)
    n_in_folders = sum(len(v) for v in _inside.values())
    n_outside = sum(len(v) for v in _front.values())

    st.markdown(
        f"""
        <div class="metric-strip">
            <div class="metric-box"><strong>{len(uploaded_files)}</strong><span>file(s) uploaded</span></div>
            <div class="metric-box"><strong>{len(_group_names)}</strong><span>combined PDF(s) to create</span></div>
            <div class="metric-box"><strong>{n_in_folders}</strong><span>file(s) inside folders</span></div>
            <div class="metric-box"><strong>{n_outside}</strong><span>outside file(s) (added to front)</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.button("Build Combined PDFs", type="primary", use_container_width=True):
        with st.spinner("Sorting files and building combined PDFs..."):
            try:
                zip_bytes, results, log, unmatched = build_account_final_pdfs(uploaded_files)
                st.session_state["docbuild_zip_bytes"] = zip_bytes
                st.session_state["docbuild_results"] = results
                st.session_state["docbuild_log"] = log
                st.session_state["docbuild_unmatched"] = unmatched
            except Exception as exc:
                st.session_state["docbuild_zip_bytes"] = None
                st.error(f"Could not build the combined PDFs: {exc}")

    results = st.session_state.get("docbuild_results") or []
    if results:
        st.subheader("Results by Account")
        st.dataframe(results, use_container_width=True, hide_index=True)

    unmatched = st.session_state.get("docbuild_unmatched") or []
    if unmatched:
        with st.expander(f"Files with no account number ({len(unmatched)}) - not included", expanded=False):
            for name in unmatched:
                st.markdown(f"- {name}")

    log = st.session_state.get("docbuild_log") or []
    if log:
        with st.expander(f"Processing log ({len(log)} note(s))", expanded=False):
            for entry in log:
                st.markdown(f"- {entry}")

    if st.session_state.get("docbuild_zip_bytes"):
        st.success("Your combined account PDFs are ready.")
        st.download_button(
            "Download Combined PDFs (ZIP)",
            data=st.session_state["docbuild_zip_bytes"],
            file_name="account_final_pdfs.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )


def get_tools() -> list[ToolDefinition]:
    return [
        ToolDefinition(
            tool_id="beam-pdf-splitter",
            name="Beam PDF Splitter",
            category="PDF",
            description="Split PDFs into two-page chunks and name each output from the beam number in the footer.",
            render=render_beam_pdf_splitter,
        ),
        ToolDefinition(
            tool_id="affidavit-pdf-splitter",
            name="Affidavit PDF Splitter",
            category="PDF",
            description="Split affidavit PDFs into two-page chunks and name each output with the OFN number plus W at the end.",
            render=render_affidavit_pdf_splitter,
        ),
        ToolDefinition(
            tool_id="spreadsheet-compare",
            name="Spreadsheet Comparison",
            category="Documents",
            description="Compare two snapshots of a spreadsheet (.xlsx or .csv). Match rows by a key column; report new, removed, and changed rows; ignore formatting-only differences.",
            render=render_spreadsheet_compare,
        ),
        ToolDefinition(
            tool_id="file-combiner",
            name="File Combiner",
            category="PDF",
            description="Sort uploaded files by account number and merge each account into one combined PDF named <account>final.pdf (letter first, then supporting documents).",
            render=render_file_combiner,
        ),
        ToolDefinition(
            tool_id="file-name-lister",
            name="File Name Lister",
            category="Documents",
            description="Upload a file or a whole folder and download a spreadsheet listing every file name, type, and size.",
            render=render_file_name_lister,
        ),
    ]


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    inject_css()

    tools = get_tools()
    menu_options = ["Home"] + [tool.name for tool in tools] + ["Add Future Tools"]
    active_label = st.session_state.get("active_tool_label", "Home")
    if active_label not in menu_options:
        active_label = "Home"

    with st.sidebar:
        st.title(APP_TITLE)
        selected_label = st.radio(
            "Tools",
            menu_options,
            index=menu_options.index(active_label),
            label_visibility="collapsed",
        )
        st.session_state["active_tool_label"] = selected_label
        st.divider()
        st.caption("Upload files, process them in the browser app, then download results.")

    if selected_label == "Home":
        render_home(tools)
        return

    if selected_label == "Add Future Tools":
        render_future_tools_guide()
        return

    selected_tool = next(tool for tool in tools if tool.name == selected_label)
    selected_tool.render()


if __name__ == "__main__":
    main()
